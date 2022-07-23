import openpyxl

def verificar_horarios(dia, hora, matr_atuais, cc_obrig, cc_elet, checagem, flag):
    
    if (checagem == 0):
        return flag

    flag1 = 0
    atual = matr_atuais[checagem - 1]
    for row in cc_obrig.values:
        for value in row:
            if (value == atual):
                tupla = row
                flag1 = 1
                break
    
    if (flag1 == 0):
        atual = matr_atuais[checagem - 1]
        for row in cc_elet.values:
            for value in row:
                if (value == atual):
                    tupla = row
                    break

    verificar = list(tupla)
    if (flag1 == 1):
        dia_v = verificar[5].split(', ')
        hora_v = verificar[6].split(' - ')
    elif (flag1 == 0):
        dia_v = verificar[4].split(', ')
        hora_v = verificar[5].split(' - ')

    for i in dia:
        for j in dia_v:
            if (i == j and hora[0] == hora_v[0]):
                flag = flag + 1
                break
        
    checagem = checagem - 1
    return verificar_horarios(dia, hora, matr_atuais, cc_obrig, cc_elet, checagem, flag)

def verificar_pre_requisito(disciplinas_pagas, disciplina, aluno_atual, flag):
    
    #Verifica se a disciplina em questão é o 'projetão' e se o aluno está no fluxo padrão do 5 período. 
    if (disciplina[0] == 'COMP382'):
        if (aluno_atual[3] == 5.0 and aluno_atual[4] == 'Padrão'):
            flag = 1
            return flag
        else:
            return flag
        
    #Verifica se a disciplina tem algum pré requisito.
    elif (disciplina[2] == 0.0):
        flag = 1
        return flag
    
    else:
        
        #Cria uma lista com os pré requisitos.
        requisitos = disciplina[2].split(', ')
        
        check = 0
        
        #Checa se as disciplinas pagas batem com os pré requisitos.
        for i in range(len(requisitos)):
            for j in range(len(disciplinas_pagas)):
                if (requisitos[i] == disciplinas_pagas[j]):
                    check = check + 1
        
        #Retorna se o aluno pode ou não efetuar a matrícula.           
        if (check == len(requisitos)):
            flag = 1
            return flag
        else:
            return flag
        
    
def pedir_disciplina (aluno_atual, cc_obrig, cc_elet, matr_atuais):
    
    print ('\n1 - Solicitar matrícula.')
    print ('2 - Sair.')
    opcao = int(input('Digite a opção desejada: '))
    
    cont = len(matr_atuais)
    
    #Verifica se o limite de disciplinas foi atingido.
    if (cont == 10 and opcao == 1):
        aluno_atual[3] = aluno_atual[3] + 1
        aluno_atual[7] = matr_atuais
        print ('\nVocê atingiu o limite máximo de 10 disciplinas para esse período.\nSessão encerrada.\n')
        return matr_atuais
    
    #Verifica se o aluno pode sair, já que precisa de ao menos 3 disciplinas.
    elif (opcao == 2 and cont < 3):
        print ('\nVocê não atingiu o mínimo de 3 disciplinas para o período.\nPor favor, continue a matrícula.')
        return pedir_disciplina(aluno_atual, cc_obrig, cc_elet, matr_atuais)

    elif (opcao == 2):
        aluno_atual[3] = aluno_atual[3] + 1
        aluno_atual[7] = matr_atuais
        return matr_atuais
    
    else:
        
        flag = 0
        
        #Cria uma lista com as disciplinas já pagas.
        disc = aluno_atual[5]
        disciplinas_pagas = disc.split(', ')
        
        verificar = input ('Digite o código da disciplina em que deseja se matricular: ')
    
        #Procura o código da disciplina informado pelo aluno nas obrigatórias de CC.
        for row in cc_obrig.values:
            for value in row:
                if (value == verificar):
                    codigo = row
                    flag = 1
                    break
        
        #Procura o código da disciplina informado pelo aluno nas eletivas de CC, caso a anterior não tenha achado.
        if (flag == 0):
            for row in cc_elet.values:
                for value in row:
                    if (value == verificar):
                        codigo = row
                        flag = 2
                        break
        
        #informa um erro de código inexistente.  
        if (flag == 0):
            print ('\nO código informado não corresponde a nenhuma disciplina.\nPor favor, tente novamente!')            
            return pedir_disciplina(aluno_atual, cc_obrig, cc_elet, matr_atuais)

        #Armazena em lista a disciplina solicitada.
        disciplina = list(codigo)

        #Chama a função de verificação.
        result_requis = verificar_pre_requisito(disciplinas_pagas, disciplina, aluno_atual, 0)
        checagem = len(matr_atuais)
        if (flag == 1):
            dia = disciplina[5].split(', ')
            hora = disciplina[6].split(' - ')
        elif (flag == 2):
            dia = disciplina[4].split(', ')
            hora = disciplina[5].split(' - ')
        result_horario = verificar_horarios(dia, hora, matr_atuais, cc_obrig, cc_elet, checagem, 0)
    
        #Condição executada caso a o aluno esteja em dia com os pré requisitos.
        if (result_requis >= 1 and result_horario == 0):
            matr_atuais.append(verificar)
            print ('\nMatrícula realizada com sucesso!')
            print ('Você já está matriculado em:\n')
            for i in matr_atuais:
                print (f'{i}')
            return pedir_disciplina(aluno_atual, cc_obrig, cc_elet, matr_atuais)
        
        #Condição executada caso o aluno não possua algum pré requisito ou haja choque de horários.
        elif (result_requis == 0):
            print ('\nMatrícula negada!\nVocê não possui os requisitos dessa disciplina.\n')
            return pedir_disciplina(aluno_atual, cc_obrig, cc_elet, matr_atuais)
        elif (result_horario > 0):
            print ('Matrícula negada!\nVocê já está matriculado em disciplina(s) nesse mesmo dia/hora,\nnão pode haver choque de horários.\n')
            return pedir_disciplina(aluno_atual, cc_obrig, cc_elet, matr_atuais)
        
#Primeiro comando a ser executado na tela.
matricula = float(input ('Digite o número de matrícula: '))

#Abre as planilhas do Excel.
path_dados_alunos = openpyxl.load_workbook("C:\\Users\\Pichau\\OneDrive\\Área de Trabalho\\ProjetoED-main\\dados_dos_alunos.xlsx")
path_disciplinas = openpyxl.load_workbook("C:\\Users\\Pichau\\OneDrive\\Área de Trabalho\\ProjetoED-main\\Disciplinas - Trabalho Estrutura de Dados.xlsx")

#Especifica e armazena as Sheets que serão utilizadas.
cc_obrig = path_disciplinas['Obrigatórias CC']
cc_elet = path_disciplinas['Eletivas CC']
pagina = path_dados_alunos['Página1']

#Procura a matrícula informada na planilha usada como 'base de dados dos alunos'.
for row in pagina.values:
    for value in row:
        if (value == matricula):
            aluno = row
            break

#Converte a tupla alunos para lista.
aluno_atual = list(aluno)

#Tela de boas vindas e chamamento da função para pedir disciplinas.
print (f'\nBem vindo(a) à página de matrícula, {aluno_atual[0]}!')
print ('\n----------DADOS DO(A) ALUNO(A)----------\n')
print (f'\nAluno: {aluno_atual[0]}.')
print (f'CPF: {aluno_atual[1]}.')
print (f'Número de matrícula: {aluno_atual[2]}.')
print (f'Período matriculado: {aluno_atual[3]}.')
print (f'Fluxo: {aluno_atual[4]}.')
print (f'Matrículas do período: {aluno_atual[7]}.')
print ('\n-----------------------------------------\n')

periodo = int(aluno_atual[3])
matr_atuais = []
matr = pedir_disciplina(aluno_atual, cc_obrig, cc_elet, matr_atuais)

#Tela com as informações atualizadas do aluno que efetuou sua matrícula.
print ('\n----------DADOS DO(A) ALUNO(A)----------\n')
print (f'\nAluno: {aluno_atual[0]}.')
print (f'CPF: {aluno_atual[1]}.')
print (f'Número de matrícula: {aluno_atual[2]}.')
print (f'Período matriculado: {aluno_atual[3]}.')
print (f'Fluxo: {aluno_atual[4]}.')
print (f'Matrículas do período: {aluno_atual[7]}.')
print ('\n-----------------------------------------\n')

#Atualiza os valores na planilha.
i = 1

for row in pagina.values:
    for value in row:
        if (value == matricula):
            pagina.cell(row=i, column=4).value = periodo + 1
            disc_ativas = ', '.join(matr)
            pagina.cell(row=i, column=8).value = disc_ativas
            break
    i = i + 1

#Salva na planilha as alterações realizadas.
path_dados_alunos.save('dados_dos_alunos.xlsx')
