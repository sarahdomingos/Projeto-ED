# Aqui ficam funções utilitárias para os outros arquivos 
import os
from openpyxl import load_workbook

def checkMatriculaExiste(numero):

    numeroExiste = False
    nome = None
    
    file = os.getcwd() + '/dados_dos_alunos.xlsx'
    wb = load_workbook(file)
    page = wb['Página1']

    for index in range(1, len(page['C'])):
        if (page['C'][index].value == numero):
            numeroExiste = True
            break
    
    if (numeroExiste == False):
        index = -1 


    return [numeroExiste, index, page['A'][index].value]

def resgatarDisciplinasAtivas(index):

    file = os.getcwd() + '/dados_dos_alunos.xlsx'
    wb = load_workbook(file)
    page = wb['Página1']

    materias = page['H'][index].value

    if (materias != None):
        materias = materias.split(', ')
    else:
        materias = []

    return materias

def removerDisciplina(codigoDisciplina, index):
    file = os.getcwd() + '/dados_dos_alunos.xlsx'
    wb = load_workbook(file)
    page = wb["Página1"]

    disciplinas = page['H'][index].value
    disciplinas = disciplinas.split(", ")

    print('Tamanho', len(disciplinas))

    for i in range(0, len(disciplinas)):
        print(disciplinas[i] == codigoDisciplina)
        if (disciplinas[i] == codigoDisciplina):
            disciplinas.pop(i)
    
    print(', '.join(disciplinas))

    page['H'][index].value = ', '.join(disciplinas)

def adicionarDisciplinas(disciplinas, index):
    file = os.getcwd() + '/dados_dos_alunos.xlsx'
    wb = load_workbook(file)
    page = wb["Página1"]
    
    page['H'][index].value = ', '.join(disciplinas)

    try: 
        # wb.save("dados_dos_alunos.xlsx")
        return True 
    except:
        return False 

def resgatarDadosDisciplinas(codigo):

    file = os.getcwd() + '/Disciplinas - Trabalho Estrutura de Dados.xlsx'
    wb = load_workbook(file)
    page = wb['Obrigatórias CC']    

    disciplina = {}

    found = False

    for index in range(1, len(page['A'])):
        if (page['A'][index].value == codigo):
            disciplina = {
                "codigo": page['A'][index].value,
                "nome": page['B'][index].value,
                # "periodo": page['D'][index].value,
                "dias": page['F'][index].value,
                "carga horária": page['E'][index].value,
                "horario": page['G'][index].value
            }
            break
        
    if (not found):
        page = wb['Eletivas CC']

        for index in range(1, len(page['A'])):
            if (page['A'][index].value == codigo):
                disciplina = {
                    "codigo": page['A'][index].value,
                    "nome": page['B'][index].value,
                    "dias": page['E'][index].value, 
                    "carga horária": page['D'][index].value,
                    "horario": page['F'][index].value
                }
                break 

    return disciplina

def resgatarMateriasPagas(index):
    file = os.getcwd() + '/dados_dos_alunos.xlsx'
    wb = load_workbook(file)
    page = wb['Página1']

    materias = page['F'][index].value

    if (materias != None):
        materias = materias.split(', ')
    else:
        materias = []

    return materias

def checarMateriaTemVagas(codigoDisciplina):
    file = os.getcwd() + '/Disciplinas - Trabalho Estrutura de Dados.xlsx'
    wb = load_workbook(file)
    found = False 
    temVagas = False

    page = wb["Obrigatórias CC"]

    for i in range(1, len(page['A'])):
        if (page['A'][i].value == codigoDisciplina):
            found = True
            print(page['H'][i].value)
            if (int(page['H'][i].value) > 0):
                temVagas = True
                page['G'][i].value = int(page['G'][i].value) - 1
            break
    
    if (not found):
        page = wb["Eletivas CC"]

        for i in range(1, len(page['A'])):
            if (page['A'][i].value == codigoDisciplina):
                found = True
                if (int(page['G'][i].value) > 0):
                    temVagas = True
                    page['G'][i].value = int(page['G'][i].value) - 1
                break
    # wb.save("Disciplinas - Trabalho Estrutura de Dados.xlsx")

    return temVagas


def resgatarMateriasPossiveisAjuste(materiasPagas, materiasSolicitadas): 
    
    file = os.getcwd() + '/Disciplinas - Trabalho Estrutura de Dados.xlsx'
    wb = load_workbook(file)

    materiasDisponiveis = []

    page = wb["Obrigatórias CC"]

    obrigatorias_primeiro_quinto = []

    for i in range(1, len(page['A'])):

        if (page['D'][i].value != None and page['D'][i].value >= 1 and page['D'][i].value <= 5):
            obrigatorias_primeiro_quinto.append(page['A'][i].value)


    obrigatorias_possiveis = []
    eletivas_possiveis = []

    for i in range(1, len(page['A'])):
        if (page['A'][i].value != None) and page['A'][i].value not in materiasPagas and page['A'][i].value not in materiasSolicitadas:
            #split pre requisitos 
            prerequisitos = str(page['C'][i].value)
            prerequisitos = prerequisitos.split(",")
            
            if (page['C'][i].value == 'TODAS AS DISCIPLINAS OBRIGATÓRIAS DO 1º AO 5º PERÍODO'):
                obrigatorias_possiveis.append([page['A'][i].value, obrigatorias_primeiro_quinto])
            else:
                obrigatorias_possiveis.append([page['A'][i].value, prerequisitos])

    page = wb["Eletivas CC"]
    for i in range(1, len(page['A'])):
        if (page['A'][i].value != None) and page['A'][i].value not in materiasPagas and page['A'][i].value not in materiasSolicitadas:

            #split pre requisitos 
            prerequisitos = str(page['C'][i].value)
            prerequisitos = prerequisitos.split(",")

            if (page['C'][i].value == 'TODAS AS DISCIPLINAS OBRIGATÓRIAS DO 1º AO 5º PERÍODO'):
                eletivas_possiveis.append([page['A'][i].value, obrigatorias_primeiro_quinto])
            else:
                eletivas_possiveis.append([page['A'][i].value, prerequisitos])
    
    
    # 19218721

    for i in obrigatorias_possiveis:
        possivel = True 
        for valor in i[1]:
            if (valor not in materiasPagas and valor not in materiasSolicitadas):
                possivel = False
        if (possivel):
            materiasDisponiveis.append(i[0])

    for i in eletivas_possiveis:
        possivel = True 
        for valor in i[1]:
            if (valor not in materiasPagas and valor not in materiasSolicitadas):
                possivel = False
        
        if (possivel):
            materiasDisponiveis.append(i[0])
        

    for i in range(0, len(materiasDisponiveis)):
        materiasDisponiveis[i] = resgatarDadosDisciplinas(materiasDisponiveis[i])

    return materiasDisponiveis

def resgatarMateriasPossiveisReajuste(materiasPagas): 

    file = os.getcwd() + '/Disciplinas - Trabalho Estrutura de Dados.xlsx'
    wb = load_workbook(file)

    disciplinas = []

    page = wb["Obrigatórias EC"]
    for index in range(1, len(page['A'])):
        # checa se a disciplina tem pre-requisitos
        if (page['C'][index].value == 0):
            disciplinas.append({
                "codigo": page['A'][index].value,
                "nome": page['B'][index].value,
                "carga horária": page['D'][index].value,
                "dias": page['E'][index].value,
                "horarios": page['F'][index].value
            })

    page = wb["Eletivas EC"]
    for index in range(1, len(page['A'])):
        if (page['C'][index].value == 0):
            disciplinas.append({
                "codigo": page['A'][index].value,
                "nome": page['B'][index].value,
                "carga horária": page['D'][index].value,
                "dias": page['E'][index].value,
                "horarios": page['F'][index].value
            })

    page = wb["IM"]
    for index in range(1, len(page['A'])):
        if (page['C'][index].value == 0):
            disciplinas.append({
                "codigo": page['A'][index].value,
                "nome": page['B'][index].value,
                "carga horária": page['D'][index].value,
                "dias": page['E'][index].value,
                "horarios": page['F'][index].value
            })

    page = wb["IF"]
    for index in range(1, len(page['A'])):
        if (page['C'][index].value == 0):
            disciplinas.append({
                "codigo": page['A'][index].value,
                "nome": page['B'][index].value,
                "carga horária": page['D'][index].value,
                "dias": page['E'][index].value,
                "horarios": page['F'][index].value
            })

    return disciplinas

def dadosAluno(matricula):

    file = os.getcwd() + '\dados_dos_alunos.xlsx'
    wb = load_workbook(file)
    pagina = wb['Página1']

    aluno = []

    for row in pagina.values:
        for value in row:
            if (value == matricula):
                aluno = list(row) 
                break

    return aluno

def prioridadesMatricula(matricula):

    aluno = dadosAluno(matricula)

    if aluno[3] == 1:
        prioridade = 1
    elif aluno[4] == 'Padrão' and aluno[3] >= 8:
        prioridade = 3
    elif aluno[4] == 'Padrão':
        prioridade = 0
    elif aluno[4] == 'Individual':
        prioridade = 2

    return prioridade

def confirmar(pedidos):

    file2 = os.getcwd() + '\Disciplinas - Trabalho Estrutura de Dados.xlsx'
    path_disciplinas = load_workbook(file2)

    cc_obrig = path_disciplinas['Obrigatórias CC']
    cc_elet = path_disciplinas['Eletivas CC']

    listaFinal = {}

    for i in sorted(pedidos, key=pedidos.get):
        listaFinal[i] = pedidos[i]

    for i in listaFinal:
        aluno = dadosAluno(i)
        discPedidas =  aluno[7].split(', ')
        matricula = str(aluno[2])
        for j in discPedidas:
            k = 1
            flag = 0

            for row in cc_obrig.values:
                for value in row:
                    if (value == j):
                        codigo = row
                        flag = 1
                        break
            if (flag == 0):
                for row in cc_elet.values:
                    for value in row:
                        if (value == j):
                            codigo = row
                            flag = 2
                            break

            disciplina = list(codigo)
            if flag == 1:
                vaga = int(disciplina[7])
            if flag == 2:
                vaga1 = int(disciplina[6])

            if vaga > 0 and flag == 1:
                for row in cc_obrig.values:
                    for value in row:
                        if (value == j):
                            cc_obrig.cell(row=k, column=8).value = vaga - 1
                            turma = cc_obrig.cell(row=k, column=9).value
                            if turma == 'Vazia':
                                cc_obrig.cell(row=k, column=9).value = matricula
                            else:
                                pessoas = turma.split(', ')
                                pessoas.append(matricula)
                                turmaFinal = ', '.join(pessoas)
                                cc_obrig.cell(row=k, column=9).value = turmaFinal
                            break
                    k = k + 1
            elif vaga1 > 0 and flag == 2:
                for row in cc_elet.values:
                    for value in row:
                        if (value == j):
                            cc_elet.cell(row=k, column=7).value = vaga1 - 1
                            turma = cc_elet.cell(row=k, column=8).value
                            if turma == 'Vazia':
                                cc_elet.cell(row=k, column=8).value = matricula
                            else:
                                pessoas = turma.split(', ')
                                pessoas.append(matricula)
                                turmaFinal = ', '.join(pessoas)
                                cc_elet.cell(row=k, column=8).value = turmaFinal
                            break
                    k = k + 1
            else: 
                print(f'As vagas para {disciplina[1]} já foram preenchidas!')

    path_disciplinas.save('Disciplinas - Trabalho Estrutura de Dados.xlsx')

    return listaFinal 













