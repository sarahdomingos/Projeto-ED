import random
from utils import checkMatriculaExiste

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

#Esta função é chamada quando o usuário apresenta uma entrada inválida
def tente_novamente(inicio):

    print("------------------------------------------------\n")
    print("Gostaria de tentar novamente?\n")
    print("1 - Sim")
    print("2 - Não\n")

    opcao = input("Digite o número correspondente a sua resposta: ")

    #Aqui estamos dizendo que uma entrada vazia pode ser interpretada como um espaço vazio
    if opcao == '':

        opcao = ' '

    #Aqui estamos verificando se a entrada dada pelo usuário é um número inteiro
    try:
        int(opcao)
        verificacao_numero_inteiro = True

    except ValueError:
        verificacao_numero_inteiro = False

    #Se a entrada for um número inteiro, então...
    if verificacao_numero_inteiro == True:

        #Transformamos a entrada que estaa no formato de escrita em numeral
        opcaoConvert = int(opcao)

        #Se o usuário escolher a opção 1, então o mandamos de volta para o menu inicial
        if opcaoConvert == 1:

            from MenuPrincipal import menu_Principal
            menu_Principal('iniciar')

        #Se o usuário escolher a opção 2, então encerramos o programa
        if opcaoConvert == 2:

            print("------------------------------------------")
            print("Programa encerrado.")
            print("------------------------------------------")

        #Se a entrada não for nem 1 e nem 2, então chamamos novamente a função de erro "tente_novamente"
        if opcaoConvert != 1 and opcaoConvert != 2:

            print("*********************************\n")
            print("Entrada Inválida!")
            print("Você só pode digitar 1 ou 2.")
            print("*********************************\n")

            tente_novamente('iniciar')

    #Se a entrada não for um número inteiro, então chamamos a função de erro 'tente_novamente'
    if verificacao_numero_inteiro == False:

            print("*********************************\n")
            print("Entrada Inválida!")
            print("Você só pode digitar 1 ou 2.\n")
            print("*********************************\n")

            tente_novamente('iniciar')


def matricular_calouros(inicio):

    print("\n")
    print("####################################################\n")
    print("Para prosseguir com o cadastro, por favor insira os dados requeridos: \n")

    nome_aluno = input("Digite seu nome completo: ")

    #Se o nome for vazio, ele é inválido
    if nome_aluno == '' or nome_aluno == " ":

        print("Nome inválido! O nome não pode ser vazio.")
        tente_novamente("iniciar")

    #Aqui verificamos se existe número na entrada
    verificacao_existe_numero = any(chr.isdigit() for chr in nome_aluno)
    
    #Se existir algum dígito, o nome é inválido
    if verificacao_existe_numero == True:

        print("Nome inválido! Você não pode ter números no nome.")
        tente_novamente("iniciar")


    cpf_aluno = input("Digite seu CPF (apenas números, sem traço ou pontos): ")

    cpf_vetor = list(cpf_aluno)

    if ',' in cpf_vetor or '.' in cpf_vetor or '-' in cpf_vetor:

        print("Entrada inválida! O CPF deve conter apenas números.")
        tente_novamente("iniciar")

    matriculaValida = False 

    while (not matriculaValida):
        matricula = random.randint(10000,20000)
        matriculaValida = checkMatriculaExiste(matricula)

    curso = "Ciência da Computação"

    # Aqui estamos acessando a planilha
    wb = load_workbook('dados_dos_alunos.xlsx')
    # Aqui estamos acessando a página da planilha
    ws = wb['Página1']
    # Aqui estamos adicionando os dados dos calouros na planilha e salvando o arquivo
    to_append = [nome_aluno, cpf_aluno, matricula, 1, 'Padrão', 'COMP359, COMP360, COMP361, COMP362, COMP363', 0, 'COMP359, COMP360, COMP361, COMP362, COMP363']
    ws.append(to_append)
    wb.save('dados_dos_alunos.xlsx')

    print("Matrícula efetuada com sucesso!\n")

    print("----------------------------------------------------\n")
    print("                    Dados do Estudante\n")

    print("Nome: ", nome_aluno)
    print("CPF: ",cpf_aluno)
    print("Matrícula: ",matricula,"\n")
    print("Curso: ", curso)

    print("______________________________________________________________")
    print("|          Disciplina          |       Carga Horária          |")
    print("|_____________________________________________________________|")
    print("|          Programação I       |            72                |")
    print("|_____________________________________________________________|")
    print("|Computação, Sociedade e Ética |            72                |")
    print("|_____________________________________________________________|")
    print("|    Lógica para Computação    |            72                |")
    print("|_____________________________________________________________|")
    print("|      Matemática Discreta     |            72                |")
    print("|_____________________________________________________________|")
    print("|Cálculo Integral e Diferencial|            144               |")
    print("|_____________________________________________________________|\n")

    print("Carga horária total: 360 horas")

    return matricula



